from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "UX_ISSUE_TRACKER.xlsx"


HEADERS = [
    "ID",
    "标题",
    "具体问题",
    "优先级分析",
    "建议优先级",
    "当前状态",
    "来源",
    "发现日期",
    "是否立即处理",
]

ROWS = [
    [
        "UX-20260528-001",
        "玩家 Web 端缺少设计审美",
        "当前玩家端虽然已经拆出 WebView，但视觉层面仍偏基础页面，缺少游戏客服入口应有的品牌感、空间节奏、组件精致度和移动端 WebView 质感。",
        "不阻断问答、反馈提交和进度查询，但会明显影响玩家端可信度，也会影响后续进入视觉设计阶段的输入质量。属于“产品观感与使用信任”问题。",
        "P1",
        "已处理，待评测",
        "用户反馈",
        "2026-05-28",
        "否",
    ],
    [
        "UX-20260528-002",
        "快捷入口交互预期不一致",
        "玩家端点击“玩法规则/活动奖励/充值支付/BUG举报建议”等快捷入口后，当前行为是把一个示例问题填入输入框；用户预期是 Agent 继续追问“你在这个分类下具体遇到了什么问题”，而不是随机给出示例问题。",
        "该问题会影响玩家端核心交互理解。快捷入口是玩家进入客服后的高频操作，如果行为不符合预期，会让玩家觉得系统在替自己输入问题，而不是引导表达诉求。影响核心体验但不阻断完整链路。",
        "P1",
        "已处理，待评测",
        "用户反馈",
        "2026-05-28",
        "否",
    ],
    [
        "UX-20260528-003",
        "玩家输入需要增加模糊词理解",
        "玩家 Web 端对玩家输入的自然语言还不够“游戏化”。玩家可能会把游戏内货币说成“金币”“豆”“欢乐豆”，这几类表达应统一理解为游戏内金币。玩家说“钱”时，需要结合上下文判断是游戏内金币，还是现实人民币/充值金额。",
        "这会影响意图识别、实体抽取和反馈分类准确性。尤其是“钱没到账”“钱少了”这类表达，既可能是充值人民币问题，也可能是游戏内金币/欢乐豆结算问题。如果不处理，容易把普通游戏币问题误判为支付充值 P0，或把支付问题误判为游戏内结算问题。它影响核心 Agent 判断质量，建议标为 P1。",
        "P1",
        "待评估，暂不处理",
        "用户反馈",
        "2026-05-28",
        "否",
    ],
    [
        "UX-20260528-004",
        "Agent 回复缺少人情味",
        "当前 Agent 回复偏功能性，能回答和追问，但语气较机械。对于充值不到账、奖励未到账、卡死、举报等有明显负面情绪的场景，需要增加更有安抚感和服务感的话术，例如先表达理解，再说明会帮玩家整理信息并提交核查。",
        "该问题不影响链路可用性，但会影响客服体验和玩家信任感。尤其游戏客服场景里，玩家经常带着不满情绪进入客服入口，回复缺少人情味会削弱“智能客服”的真实感。建议标为 P2，排入话术优化。",
        "P2",
        "待评估，暂不处理",
        "用户反馈",
        "2026-05-28",
        "否",
    ],
]


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I3"
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="145C54")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E0E8"))
    wrap = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 18,
        "B": 24,
        "C": 46,
        "D": 46,
        "E": 12,
        "F": 18,
        "G": 12,
        "H": 14,
        "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap
            cell.border = border
        row[4].fill = PatternFill("solid", fgColor="FEF0C7")
        row[4].font = Font(color="A15C07", bold=True)
        row[5].fill = PatternFill("solid", fgColor="F0F3F7")
        row[8].fill = PatternFill("solid", fgColor="E4F3F1")

    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 96

    table = Table(displayName="UXIssueTracker", ref=f"A1:I{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "体验问题台账"
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)
    style_sheet(ws)

    summary = wb.create_sheet("说明")
    summary["A1"] = "体验问题台账说明"
    summary["A1"].font = Font(size=16, bold=True, color="145C54")
    summary["A3"] = "用途"
    summary["B3"] = "记录开发过程中发现的体验问题、影响分析和处理优先级；只记录，不代表立即处理。"
    summary["A4"] = "当前问题数"
    summary["B4"] = len(ROWS)
    summary["A5"] = "使用方式"
    summary["B5"] = "优先看“体验问题台账”工作表；CSV 仅用于后续导入飞书多维表格。"
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 88
    summary.sheet_view.showGridLines = False
    for row in summary.iter_rows(min_row=1, max_row=5, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT)


if __name__ == "__main__":
    build()
    print(OUTPUT)
